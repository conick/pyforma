import warnings
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.workbook import Workbook
from openpyxl.workbook.protection import WorkbookProtection
# from pydantic import BaseModel
from pydantic.dataclasses import dataclass
from typing import Optional, Union


@dataclass
class ExcelCellInterval():
    row_start: int
    row_end: int
    col_start: int
    col_end: int

class ExcelColumnInterval():
    start: int
    end: int


class ExcelManager():
    _file_path: str = None
    _wb: Workbook = None

    def __init__(self, file_path: str) -> None:
        self._file_path = file_path
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            self._wb = load_workbook(file_path)

    def lock(
        self,
        unlock_columns: Optional[list[str]],
        # unlock_cells: Optional[list[ExcelCellInterval]] = None
    ):
        self._wb.security = WorkbookProtection(lockStructure=True, lockWindows = True)
        ws = self._wb.active
        prot = ws.protection
        prot.enabled = True
        prot.selectLockedCells = False
        prot.selectUnlockedCells = False
        prot.algorithmName = None
        prot.sheet = True
        prot.objects = False
        prot.insertRows = True
        prot.insertHyperlinks = True
        prot.autoFilter = False
        prot.scenarios = False
        prot.formatColumns = True
        prot.deleteColumns = True
        prot.insertColumns = True
        prot.pivotTables = True
        prot.deleteRows = True
        prot.formatCells = True
        prot.saltValue = None
        prot.formatRows = True
        prot.sort = False
        prot.spinCount = None
        prot.hashValue = None

        if unlock_columns:
            for col in unlock_columns:
                for cell in ws[col]:
                    cell.protection = Protection(locked=False)
        return

        # BELOW CODE DIDN'T WORK AS EXPECTED! 
        # It locks 10 columns even if we choose only one
        # So this feature is currrently disabled

        # for iv in unlock_cells:
        #     for col in range(iv.col_start, iv.col_end + 1):
        #         for row in range(iv.row_start, iv.row_end):
        #             ws.cell(row = x, column = y).protection = Protection(locked=False, hidden=False)
        #             # print(f'lock row - {x} column - {y}')
        #             # ws.cell(row = row, column = col).value = ' hello '
    
    def auto_filter(self):
        ws = self._wb.active
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.enable = True

    def save(self):
        self._wb.save(self._file_path)
